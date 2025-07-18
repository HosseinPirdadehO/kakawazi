from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from .models import User, PhoneOTP, Referral
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse


# =================== Inline ===================
class ReferralInline(admin.TabularInline):
    model = Referral
    fk_name = 'inviter'
    extra = 0
    readonly_fields = ('invited', 'created_at')
    verbose_name = "کاربر دعوت‌شده"
    verbose_name_plural = "کاربران دعوت‌شده"


# =================== UserAdmin ===================
@admin.register(User)
class UserAdmin(BaseUserAdmin):
    inlines = [ReferralInline]

    fieldsets = (
        ("اطلاعات ورود", {
            'fields': ('phone_number', 'password', 'is_phone_verified')
        }),
        ("اطلاعات شخصی", {
            'fields': ('first_name', 'last_name', 'email', 'birth_date', 'image', 'school')
        }),
        ("موقعیت جغرافیایی", {
            'fields': ('state', 'city', )
        }),
        ("اطلاعات ماشین", {
            'fields': ('plate_number', 'type_of_car')
        }),
        ("نقش و وضعیت", {
            'fields': ('system_role', 'job_role', 'status', 'is_active', 'is_staff', 'is_superuser')
        }),
        ("سیستم ارجاع", {
            'fields': ('referral_code', 'referral')
        }),
        ("زمان‌ها", {
            'fields': ('created_at', 'date_joined', 'last_login')
        }),
        ("دسترسی‌ها", {
            'fields': ('groups', 'user_permissions')
        }),
    )

    add_fieldsets = (
        ("ساخت کاربر جدید", {
            'classes': ('wide',),
            'fields': (
                'phone_number',
                'password1',
                'password2',
                'system_role',
                'school',
                'job_role',
                'is_staff',
                'is_superuser',
                'is_active',
            ),
        }),
    )

    list_display = (
        'phone_number',
        'get_full_name',
        'system_role',
        'job_role',
        'referral_link',
        'school',
        'referral_count',
        'is_staff',
        'is_active',
    )

    list_filter = (
        'system_role',
        'job_role',
        'is_staff',
        'is_active',
        'status',
    )

    search_fields = (
        'phone_number',
        'first_name',
        'last_name',
        'email',
        'referral_code',
    )

    ordering = ('-created_at',)

    readonly_fields = ('created_at', 'date_joined',
                       'last_login', 'referral_code')

    # ================ اکشن XLSX =================
    actions = ["export_all_fields_as_xlsx"]

    def export_all_fields_as_xlsx(self, request, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Users"

        # گرفتن همه فیلدهای مدل
        model = self.model
        field_names = [field.name for field in model._meta.get_fields()
                       if not field.many_to_many and not field.one_to_many]

        # نوشتن عنوان‌ها
        bold_font = Font(bold=True)
        for col_num, column_title in enumerate(field_names, 1):
            cell = worksheet.cell(row=1, column=col_num, value=column_title)
            cell.font = bold_font

        # نوشتن داده‌ها
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(field_names, 1):
                try:
                    value = getattr(obj, field, '')
                    related_field = model._meta.get_field(field)

                    # ForeignKey ها
                    if related_field.is_relation and value:
                        value = str(value)
                    # تاریخ‌ها
                    elif hasattr(value, 'strftime'):
                        value = value.strftime("%Y-%m-%d %H:%M")
                    # بقیه موارد (UUID, Decimal, JSON و ...)
                    elif not isinstance(value, (str, int, float, bool, type(None))):
                        value = str(value)
                except Exception:
                    value = ""

                worksheet.cell(row=row_num, column=col_num, value=value)

        # پاسخ HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=users_full.xlsx'
        workbook.save(response)
        return response

    export_all_fields_as_xlsx.short_description = "دانلود کامل XLSX کاربران"

    # ================ Helper methods =================
    def get_full_name(self, obj):
        return obj.get_full_name() or "—"
    get_full_name.short_description = 'نام کامل'

    def referral_link(self, obj):
        if obj.referral:
            return obj.referral.get_full_name() or obj.referral.phone_number
        return "-"
    referral_link.short_description = "دعوت‌کننده"

    def referral_count(self, obj):
        return obj.referrals.count()
    referral_count.short_description = "تعداد ارجاعات"


# =================== PhoneOTPAdmin ===================
@admin.register(PhoneOTP)
class PhoneOTPAdmin(admin.ModelAdmin):
    list_display = (
        'phone_number', 'purpose', 'created_at', 'is_verified', 'failed_attempts', 'is_expired_display'
    )
    list_filter = ('purpose', 'is_verified', 'created_at')
    search_fields = ('phone_number',)
    readonly_fields = ('code', 'created_at')

    def is_expired_display(self, obj):
        return obj.is_expired()
    is_expired_display.boolean = True
    is_expired_display.short_description = "منقضی؟"


# =================== ReferralAdmin ===================
@admin.register(Referral)
class ReferralAdmin(admin.ModelAdmin):
    list_display = ('inviter_name', 'invited_name', 'created_at')
    search_fields = ('inviter__phone_number', 'invited__phone_number')
    list_filter = ('created_at',)
    readonly_fields = ('inviter', 'invited', 'created_at')

    def inviter_name(self, obj):
        return obj.inviter.get_full_name() or obj.inviter.phone_number
    inviter_name.short_description = "دعوت‌کننده"

    def invited_name(self, obj):
        return obj.invited.get_full_name() or obj.invited.phone_number
    invited_name.short_description = "دعوت‌شده"
